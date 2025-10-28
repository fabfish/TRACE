import pandas as pd
import numpy as np
import os
import re
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font
import dataframe_image as dfi
import matplotlib.pyplot as plt
import ast
import io
import PIL.Image

# --- 1. 定义文件路径和标签 ---
# ⚠️ 警告：请确保您已将这些路径替换为正确的本地路径
file_paths = {
    "1L": "/data/yuzhiyuan/outputs_LLM-CL/Llama-3.2-1B-Instruct/cl/upcycle/newnew/predictions_single/evaluation_matrix.xlsx",
    "1L_G": "/data/yuzhiyuan/outputs_LLM-CL/Llama-3.2-1B-Instruct/cl/upcycle/grouped_1/evaluation_matrix.xlsx",
    "2L": "/data/yuzhiyuan/outputs_LLM-CL/Llama-3.2-1B-Instruct/cl/upcycle/even_new_stable/predictions_single/evaluation_matrix.xlsx",
    "2L_G": "/data/yuzhiyuan/outputs_LLM-CL/Llama-3.2-1B-Instruct/cl/upcycle/grouped_2/evaluation_matrix.xlsx",
    "Full_2L": "/data/yuzhiyuan/outputs_LLM-CL/Llama-3.2-1B-Instruct/cl/upcycle/even_full/predictions_single/evaluation_matrix.xlsx",
    "4L": "/data/yuzhiyuan/outputs_LLM-CL/Llama-3.2-1B-Instruct/cl/upcycle/four_new/predictions_single/evaluation_matrix.xlsx",
    "4L_R": "/data/yuzhiyuan/outputs_LLM-CL/Llama-3.2-1B-Instruct/cl/upcycle/four copy?/predictions_single/evaluation_matrix.xlsx",
    "Sub SFT": "/data/yuzhiyuan/outputs_LLM-CL/naive_llama3_1B_500/predictions_single/evaluation_matrix.xlsx",
    "Full SFT": "/data/yuzhiyuan/outputs_LLM-CL/naive_llama3_1B_full/predictions/evaluation_matrix.xlsx",
}

output_excel_path = "merged_evaluation_matrix_aligned.xlsx"
output_image_path = "merged_evaluation_matrix_aligned_two_columns.png" # 修改输出图片名称

# 定义方法在每个 Checkpoint 单元格内的显示顺序
STANDARD_METHODS_ORDER = ["2L", "4L_nodim", "4L_dim", "Sub SFT"]
FULL_SFT_METHOD = "Full SFT"

# 1. 固定任务顺序
TASK_ORDER = ["C-STANCE", "FOMC", "MeetingBank", "Py150", "ScienceQA", "NumGLUE-cm", "NumGLUE-ds", "20Minuten"]

# 2. 方法颜色映射
METHOD_COLORS = {
    "1L": "#FF6666",       # 浅红
    "1L_G": "#FF9999",     # 更浅红
    "2L": "#FFB266",       # 橙
    "2L_G": "#FFD966",     # 浅橙
    "Full_2L": "#FFFF66",  # 黄
    "4L": "#90EE90",       # 浅绿
    "4L_R": "#66FF66",     # 亮绿
    "Sub SFT": "#ADD8E6",  # 浅蓝
    "Full SFT": "#DDA0DD", # 紫红 (Plum)
}

METHOD_FONT_COLORS = {
    "1L": "#D00000",       # 深红
    "1L_G": "#B30000",     # 更深红
    "2L": "#E67E00",       # 深橙
    "2L_G": "#C08B00",     # 赭黄
    "Full_2L": "#808000",  # 橄榄
    "4L": "#006400",       # 深绿
    "4L_R": "#005000",     # 更深绿
    "Sub SFT": "#00008B",  # 深蓝
    "Full SFT": "#8B008B", # 深洋红 (Magenta)
}

ROW_HEIGHT = 120

def parse_metrics(metric_str):
    metrics = {}
    if pd.isna(metric_str) or not isinstance(metric_str, str):
        return metrics
    lines = metric_str.strip().split('\n')
    for line in lines:
        if ':' in line:
            tag, value = line.split(':', 1)
            metrics[tag.strip()] = value.strip()
    return metrics

def extract_sari_number(value):
    if isinstance(value, (int, float)):
        return f"{float(value):.2f}"
    if isinstance(value, str):
        match = re.search(r"['\"]?sari['\"]?\s*[:=]\s*([0-9.]+)", value)
        if match:
            return f"{float(match.group(1)):.2f}"
        try:
            return f"{float(value):.2f}"
        except Exception:
            pass
    return value

def parse_numeric_value(value_str, metric_tag):
    numeric_val = np.nan
    tag_lower = str(metric_tag).lower()

    if isinstance(value_str, (int, float)):
        numeric_val = float(value_str)
    elif isinstance(value_str, str):
        match_sari = re.search(r"['\"]?sari['\"]?\s*[:=]\s*([0-9.]+)", value_str)
        if match_sari:
            numeric_val = float(match_sari.group(1))
        else:
            try:
                numeric_val = float(value_str)
            except (ValueError, TypeError):
                numeric_val = np.nan
    
    if pd.notna(numeric_val):
        if 'sari' in tag_lower or 'similarity' in tag_lower:
            numeric_val = numeric_val / 100.0
            
    return numeric_val

def load_and_preprocess_data(file_paths):
    all_data_list = []
    for method_tag, path in file_paths.items():
        try:
            if path.endswith('.csv'):
                df = pd.read_csv(path)
            elif path.endswith('.xlsx'):
                df = pd.read_excel(path, sheet_name=0)
            else:
                continue
            df = df.rename(columns={'task_name': 'Task'})
            df = df.set_index('Task')
            for task_name in df.index:
                for col in df.columns:
                    checkpoint = str(col)
                    original_metric_str = df.loc[task_name, col]
                    metrics = parse_metrics(original_metric_str)
                    if not metrics and pd.notna(original_metric_str) and original_metric_str.strip() != "":
                        continue
                    for metric_tag, metric_value in metrics.items():
                        all_data_list.append({
                            'Method': method_tag,
                            'Task': task_name,
                            'Checkpoint': checkpoint,
                            'Metric_Tag': metric_tag,
                            'Metric_Value': metric_value
                        })
        except FileNotFoundError:
            print(f"Error: File not found at {path}")
        except Exception as e:
            print(f"Error processing file {path} for tag {method_tag}: {e}")
    final_df = pd.DataFrame(all_data_list)
    return final_df

def export_to_excel_with_merge(df, output_path):
    all_tasks = sorted(df['Task'].unique())
    all_checkpoints = sorted(
        [cp for cp in df['Checkpoint'].unique() if cp.isdigit() and int(cp) >= 0],
        key=int
    )
    all_methods = [m for m in file_paths.keys()]
    wb = Workbook()
    ws = wb.active
    ws.title = "Merged_Evaluation"
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True, size=10)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.cell(row=1, column=1, value="Epoch").font = header_font
    ws.cell(row=1, column=1).alignment = center_alignment
    current_col = 2
    for task in all_tasks:
        ws.cell(row=1, column=current_col, value=task).font = header_font
        ws.cell(row=1, column=current_col).alignment = center_alignment
        current_col += 1
    current_row = 2
    for checkpoint in all_checkpoints:
        ws.cell(row=current_row, column=1, value=f"Epoch {checkpoint}")
        ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=1).alignment = center_alignment
        for t_idx, task in enumerate(all_tasks):
            cell_content_lines = []
            for method in all_methods:
                sub_df = df[
                    (df['Checkpoint'] == str(checkpoint)) &
                    (df['Task'] == task) &
                    (df['Method'] == method)
                ]
                for _, row in sub_df.iterrows():
                    cell_content_lines.append(f"{method}_{row['Metric_Tag']}: {row['Metric_Value']}")
            ws.cell(row=current_row, column=2 + t_idx, value="\n".join(cell_content_lines))
            ws.cell(row=current_row, column=2 + t_idx).alignment = left_alignment
        current_row += 1
    for row in range(1, current_row):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.border == Border():
                cell.border = thin_border
    ws.row_dimensions[1].height = 25
    for row in range(2, current_row):
        ws.row_dimensions[row].height = ROW_HEIGHT
    for i in range(len(all_tasks) + 1):
        ws.column_dimensions[get_column_letter(i + 1)].width = 25
    wb.save(output_path)
    return current_row - 1

def is_main_metric(metric_tag, task=None):
    tag = metric_tag.lower()
    if task == "MeetingBank" and tag == "rouge-l":
        return True
    return not (('bleu' in tag) or ('rouge' in tag))

# 样式定义 (从 export_to_image 中提取出来，以便于在两个 Styler 中复用)
base_styles = [
    {'selector': 'td, th.row_heading', 'props': [
        ('vertical-align', 'middle'), 
        ('padding', '3px 5px'),
        ('font-size', '9pt'),
        ('height', '25px'),
        ('border', '1px solid #000000'), 
        ('white-space', 'pre-line')
    ]},
    {'selector': 'th.col_heading, th:not(.row_heading)', 'props': [ 
        ('text-align', 'center'), 
        ('vertical-align', 'middle'), 
        ('font-weight', 'bold'), 
        ('background-color', '#f0f0f0'),
        ('border', '1px solid #000000')
    ]},
]

# ⬇️⬇️⬇️ ** 样式函数已修改以接受 df_numeric 作为参数 ** ⬇️⬇️⬇️
def apply_all_styles(data_display, current_df_numeric):
    # data_display 是 df_display 的一个子集
    # current_df_numeric 是 df_numeric 的相应子集
    
    # 1. 提取 Full SFT 的数值 (仅限于当前处理的 Epoch 范围)
    try:
        sft_data_numeric = current_df_numeric.loc[pd.IndexSlice[:, 'Full SFT'], :]
        sft_data_numeric = sft_data_numeric.reset_index(level='Method', drop=True)
    except KeyError:
        sft_data_numeric = pd.DataFrame(columns=current_df_numeric.columns)

    style_df = pd.DataFrame('', index=data_display.index, columns=data_display.columns)

    for (epoch, method), row in current_df_numeric.iterrows():
        try:
            sft_epoch_row = sft_data_numeric.loc[epoch]
        except KeyError:
            sft_epoch_row = pd.Series(np.nan, index=current_df_numeric.columns)

        for task in row.index:
            cell_val_numeric = row[task]
            sft_val_numeric = sft_epoch_row[task]
            
            style = "" 
            
            if method == "Full SFT":
                style = (
                    f'background-color: {METHOD_COLORS.get("Full SFT", "#FFFFFF")}; '
                    f'color: #000000;'
                )
            else:
                is_better = False
                if pd.notna(cell_val_numeric) and pd.notna(sft_val_numeric):
                    is_better = cell_val_numeric > sft_val_numeric 
                
                if is_better:
                    style = (
                        f'background-color: {METHOD_COLORS.get(method, "#FFFFFF")}; '
                        f'color: #000000;'
                    )
                else:
                    style = (
                        f'background-color: #FFFFFF; '
                        f'color: {METHOD_FONT_COLORS.get(method, "#000000")};'
                    )
            
            if task == 'Avg.':
                style += '; text-align: center; font-weight: bold;'
            else:
                style += '; text-align: left;'
            
            style_df.loc[(epoch, method), task] = style

    epochs = data_display.index.get_level_values('Epoch')
    epochs_series = pd.Series(epochs.values, index=data_display.index) 
    epoch_ends = (epochs_series != epochs_series.shift(-1))
    
    border_style = 'border-bottom: 2px solid #000000;'
    
    for idx, is_end in epoch_ends.items(): 
        if is_end:
            for col in style_df.columns: 
                style_df.loc[idx, col] += f'; {border_style}'
                    
    return style_df
# ⬆️⬆️⬆️ ** 修改结束 ** ⬆️⬆️⬆️

# ⬇️⬇️⬇️ ** export_to_image 函数的大幅修改 ** ⬇️⬇️⬇️
def export_to_image(df, output_path):
    all_tasks = TASK_ORDER 
    all_checkpoints = sorted(
        [cp for cp in df['Checkpoint'].unique() if cp.isdigit() and int(cp) >= 0],
        key=int
    )
    all_methods = [m for m in file_paths.keys()] 

    # 1. 构建颜色图例 (不变)
    legend_labels = list(METHOD_COLORS.keys())
    legend_colors = [METHOD_COLORS.get(k, "#FFFFFF") for k in legend_labels]
    legend_df = pd.DataFrame([legend_labels], columns=legend_labels)
    legend_styler = legend_df.style.set_table_styles(base_styles) # 使用 base_styles
    legend_styler = legend_styler.apply(lambda x: [f'background-color: {METHOD_COLORS.get(col, "#FFFFFF")};' for col in x.index], axis=1)
    legend_styler.set_table_attributes('style="border-collapse: collapse"')
    legend_styler.hide(axis="index")

    # 2. 构建主表格 (data_rows, numeric_rows, index_tuples)
    data_rows = []
    numeric_rows = [] 
    index_tuples = [] 

    for checkpoint in all_checkpoints:
        for method in all_methods:
            row_data = {}
            row_numeric_data = {} 
            has_data_for_this_method = False
            
            for task in all_tasks:
                cell_lines = []
                numeric_val = np.nan 
                
                sub_df = df[
                    (df['Checkpoint'] == str(checkpoint)) &
                    (df['Task'] == task) &
                    (df['Method'] == method)
                ]
                
                if not sub_df.empty:
                    has_data_for_this_method = True
                    main_metric_row = None
                    for _, row in sub_df.iterrows():
                        if is_main_metric(row['Metric_Tag'], task):
                            main_metric_row = row
                            break 
                    
                    if main_metric_row is not None:
                        tag = main_metric_row['Metric_Tag']
                        value = main_metric_row['Metric_Value']
                        
                        numeric_val = parse_numeric_value(value, tag)
                        
                        if tag.lower() == "sari":
                            display_val = float(extract_sari_number(value)) # 用于显示
                            value_fmt = f"sari: {display_val:.2f}"
                            cell_lines.append(value_fmt)
                        else:
                            try:
                                # 对于显示，不应该除以100，使用原始值
                                display_val = parse_numeric_value(value, "other") 
                                value_fmt = f"{float(display_val):.2f}"
                            except Exception:
                                value_fmt = value
                            cell_lines.append(f"{tag}: {value_fmt}")
                
                row_data[task] = "\n".join(cell_lines)
                row_numeric_data[task] = numeric_val 

            if has_data_for_this_method:
                index_tuples.append((f"Epoch {checkpoint}", method))
                data_rows.append(row_data)
                numeric_rows.append(row_numeric_data) 

    if not index_tuples:
        print("警告: 没有数据可用于生成图片。")
        return

    multi_index = pd.MultiIndex.from_tuples(index_tuples, names=["Epoch", "Method"])
    df_display = pd.DataFrame(data_rows, index=multi_index, columns=all_tasks)
    df_numeric = pd.DataFrame(numeric_rows, index=multi_index, columns=all_tasks)

    # 计算平均分 (基于 df_numeric，其中 sari 和 similarity 已 /100)
    df_numeric['Avg.'] = df_numeric[all_tasks].mean(axis=1, skipna=True) # 只对任务列求平均
    df_display['Avg.'] = df_numeric['Avg.'].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "")

    # 3. 拆分 DataFrame 为两栏
    # 找到 Epoch 的分割点
    epochs_in_data = sorted(list(set([int(re.search(r'\d+', e).group()) for e in df_display.index.get_level_values('Epoch')])))
    
    first_col_epochs = [f"Epoch {e}" for e in range(4)] # 0, 1, 2, 3
    second_col_epochs = [f"Epoch {e}" for e in range(4, 8)] # 4, 5, 6, 7

    # 筛选出第一栏和第二栏的数据
    df_display_col1 = df_display.loc[df_display.index.get_level_values('Epoch').isin(first_col_epochs)]
    df_numeric_col1 = df_numeric.loc[df_numeric.index.get_level_values('Epoch').isin(first_col_epochs)]

    df_display_col2 = df_display.loc[df_display.index.get_level_values('Epoch').isin(second_col_epochs)]
    df_numeric_col2 = df_numeric.loc[df_numeric.index.get_level_values('Epoch').isin(second_col_epochs)]

    # 4. 为每栏创建 Styler
    styler_col1 = df_display_col1.style.set_table_styles(base_styles)
    styler_col1.set_table_attributes('style="border-collapse: collapse"')
    styler_col1 = styler_col1.apply(lambda x: apply_all_styles(df_display_col1, df_numeric_col1), axis=None)

    styler_col2 = df_display_col2.style.set_table_styles(base_styles)
    styler_col2.set_table_attributes('style="border-collapse: collapse"')
    styler_col2 = styler_col2.apply(lambda x: apply_all_styles(df_display_col2, df_numeric_col2), axis=None)


    # 5. 导出和拼接图片
    buf_legend = io.BytesIO()
    buf_col1 = io.BytesIO()
    buf_col2 = io.BytesIO()
    
    # 导出图例
    dfi.export(legend_styler, buf_legend, max_rows=-1, max_cols=-1, table_conversion='matplotlib', dpi=150)
    
    # 导出两栏图片
    dfi.export(styler_col1, buf_col1, max_rows=-1, max_cols=-1, table_conversion='matplotlib', dpi=150)
    dfi.export(styler_col2, buf_col2, max_rows=-1, max_cols=-1, table_conversion='matplotlib', dpi=150)
    
    buf_legend.seek(0)
    buf_col1.seek(0)
    buf_col2.seek(0)
    
    img_legend = PIL.Image.open(buf_legend)
    img_col1 = PIL.Image.open(buf_col1)
    img_col2 = PIL.Image.open(buf_col2)

    # 计算总宽度和高度
    # 留一些边距
    H_PAD = 30 # 水平分隔符宽度
    V_PAD = 10 # 垂直分隔符宽度

    max_col_height = max(img_col1.height, img_col2.height)
    total_width = img_col1.width + H_PAD + img_col2.width
    total_height = img_legend.height + V_PAD + max_col_height

    new_img = PIL.Image.new('RGB', (total_width, total_height), (255, 255, 255))
    
    # 粘贴图例
    new_img.paste(img_legend, (0, 0)) # 图例放在顶部左侧
    
    # 粘贴两栏
    new_img.paste(img_col1, (0, img_legend.height + V_PAD))
    new_img.paste(img_col2, (img_col1.width + H_PAD, img_legend.height + V_PAD)) 
    
    new_img.save(output_path)
    print(f"\n✅ 图片已成功导出到: {output_path}")

# ⬆️⬆️⬆️ ** export_to_image 函数大幅修改结束 ** ⬆️⬆️⬆️


# --- 7. 主执行流程 ---
if __name__ == "__main__":
    
    print("🚀 开始加载和预处理数据...")
    final_df = load_and_preprocess_data(file_paths)
    
    if final_df is None or final_df.empty:
        print("❌ 数据预处理失败或数据为空，请检查文件路径和内容。")
    else:
        print("✅ 数据预处理完成。")
        
        print("📝 正在导出 Excel 文件 (复杂结构/最大行对齐) ...")
        export_to_excel_with_merge(final_df, output_excel_path)
        print(f"✅ Excel 文件已成功导出到: {output_excel_path}")
        
        print("🖼️ 正在生成图片 (两栏布局) ...")
        export_to_image(final_df, output_image_path)